"""
Airtime & Bundles Reversal Processor
=====================================
Automates cross-referencing of failed ExpressPay transactions against 
PAY_PEN_PROCESS to generate reversal files.

Output matches the exact format of the manual reversal template:
  AMOUNT | ACCOUNT TO DEBIT | DEBIT ACCOUNT NAME | ACCOUNT TO CREDIT | 
  CREDIT ACCOUNT NAME | NARRATION | SESSION ID
"""

import pandas as pd
import os
import re
import json
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XlImage
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker


# Constants
EXPRESSPAY_ACCOUNT = 1020001515537
EXPRESSPAY_NAME = "EXPRESSPAY GHANA LIMITED"
STATS_LOG_FILE = "dashboard_stats.json"


def load_expresspay_extract(file_path):
    """Load ExpressPay extract CSV file."""
    print(f"Loading ExpressPay Extract: {file_path}")
    df = pd.read_csv(file_path)
    print(f"  Total transactions: {len(df)}")
    return df


def load_pay_pen_process(file_path):
    """Load PAY_PEN_PROCESS CSV file."""
    print(f"Loading PAY_PEN_PROCESS: {file_path}")
    # Skip the first row which is just a title
    df = pd.read_csv(file_path, skiprows=1)
    print(f"  Total payment records: {len(df)}")
    return df


def filter_failed_transactions(expresspay_df):
    """
    Filter for failed transactions (FULFILMENT_TIMEOUT, Third party Timeout, Invalid Request, Insufficient Funds).
    RETRY_LIMIT_REACHED are repetitions of timeouts and are excluded.
    Also excludes any TRACE_IDs that eventually succeeded, and removes duplicates.
    """
    print("\nFiltering failed transactions...")

    failure_types = ['FULFILMENT_TIMEOUT', 'Third party Timeout', 'Invalid Request', 'Insufficient Funds']
    failed_df = expresspay_df[expresspay_df['RESULT_MSG'].isin(failure_types)].copy()
    fulfilment_count = len(failed_df[failed_df['RESULT_MSG'] == 'FULFILMENT_TIMEOUT'])
    third_party_count = len(failed_df[failed_df['RESULT_MSG'] == 'Third party Timeout'])
    invalid_count = len(failed_df[failed_df['RESULT_MSG'] == 'Invalid Request'])
    insufficient_count = len(failed_df[failed_df['RESULT_MSG'] == 'Insufficient Funds'])
    print(f"  FULFILMENT_TIMEOUT transactions: {fulfilment_count}")
    print(f"  Third party Timeout transactions: {third_party_count}")
    print(f"  Invalid Request transactions: {invalid_count}")
    print(f"  Insufficient Funds transactions: {insufficient_count}")

    # Exclude any TRACE_IDs that eventually succeeded
    successful_trace_ids = expresspay_df[
        expresspay_df['RESULT_MSG'] == 'Success'
    ]['TRACE_ID'].unique()
    truly_failed = failed_df[~failed_df['TRACE_ID'].isin(successful_trace_ids)]
    print(f"  Truly failed (never succeeded): {len(truly_failed)}")

    # Remove duplicate TRACE_IDs (keep first occurrence to avoid reversing the same transaction twice)
    before_dedup = len(truly_failed)
    truly_failed = truly_failed.drop_duplicates(subset='TRACE_ID', keep='first')
    dupes_removed = before_dedup - len(truly_failed)
    if dupes_removed > 0:
        print(f"  Duplicates removed: {dupes_removed}")
    print(f"  Unique failed transactions: {len(truly_failed)}")

    return truly_failed


def match_transactions(failed_df, pay_pen_df):
    """
    Match failed transactions with PAY_PEN_PROCESS using TRACE_ID = Narration.
    """
    print("\nMatching transactions...")

    # Clean columns for matching
    pay_pen_df = pay_pen_df.copy()
    pay_pen_df['Narration'] = pay_pen_df['Narration'].astype(str).str.strip()
    failed_df = failed_df.copy()
    failed_df['TRACE_ID'] = failed_df['TRACE_ID'].astype(str).str.strip()

    # Merge on TRACE_ID = Narration
    matched = pd.merge(
        failed_df,
        pay_pen_df,
        left_on='TRACE_ID',
        right_on='Narration',
        how='inner'
    )

    print(f"  Matched transactions: {len(matched)}")

    unmatched_count = len(failed_df) - len(matched)
    if unmatched_count > 0:
        unmatched_ids = set(failed_df['TRACE_ID']) - set(matched['TRACE_ID'])
        print(f"  WARNING: {unmatched_count} failed transactions not found in PAY_PEN_PROCESS")
        for uid in list(unmatched_ids)[:5]:
            print(f"    - {uid}")

    return matched


def build_narration(txn_date_str):
    """
    Build narration string in the format: AirtimeAndBundle dd DDMMYY Rvsd
    Uses the transaction date from the extract.
    """
    try:
        txn_date = pd.to_datetime(txn_date_str)
        date_part = txn_date.strftime('%d%m%y')
    except Exception:
        date_part = datetime.now().strftime('%d%m%y')
    return f"AirtimeAndBundle dd {date_part} Rvsd "


def create_reversal_file(matched_df, output_path, bp_matched_df=None):
    """
    Create the reversal Excel file matching the exact original template format.

    Columns:
      A: AMOUNT
      B: ACCOUNT TO DEBIT  (always ExpressPay collection account)
      C: DEBIT ACCOUNT NAME (always EXPRESSPAY GHANA LIMITED)
      D: ACCOUNT TO CREDIT (customer account from PAY_PEN)
      E: CREDIT ACCOUNT NAME (customer name from PAY_PEN)
      F: NARRATION (AirtimeAndBundle dd DDMMYY Rvsd)
      G: SESSION ID (trace ID with dashes removed)

    Sheet name: today's date e.g. "25 Mar 2026"
    bp_matched_df: optional DataFrame of BP Export failures to append after extract failures.
    """
    print(f"\nGenerating reversal file: {output_path}")

    wb = Workbook()
    ws = wb.active
    ws.title = datetime.now().strftime("%#d %b %Y")

    # Write headers
    headers = [
        "AMOUNT ",
        "ACCOUNT TO DEBIT ",
        "DEBIT ACCOUNT NAME ",
        "ACCOUNT TO CREDIT",
        "CREDIT ACCOUNT NAME ",
        "NARRATION",
        "SESSION ID"
    ]
    center_align = Alignment(horizontal='center')
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)

    # Handle Amount - could be string or numeric
    matched_df = matched_df.copy()
    if matched_df['Amount'].dtype == 'object':
        matched_df['_amount'] = matched_df['Amount'].str.replace(',', '').astype(float)
    else:
        matched_df['_amount'] = matched_df['Amount'].astype(float)

    # Sort by amount smallest to biggest
    matched_df = matched_df.sort_values('_amount', ascending=True).reset_index(drop=True)

    # Write data rows
    row_num = 2
    for idx in matched_df.index:
        amount = matched_df.loc[idx, '_amount']
        # Convert to int if it's a whole number
        if amount == int(amount):
            amount = int(amount)

        customer_account = matched_df.loc[idx, 'Acc Num1']
        # Get customer name - strip titles like MR, MS, MISS, DR, DR.
        raw_name = str(matched_df.loc[idx, 'Account1 Name']).strip()
        name = raw_name
        for title in ['MR ', 'MRS ', 'MS ', 'MISS ', 'DR. ', 'DR ']:
            if name.upper().startswith(title):
                name = name[len(title):]
                break

        txn_date = str(matched_df.loc[idx, 'TXN_DATE'])
        narration = build_narration(txn_date)
        session_id = str(matched_df.loc[idx, 'TRACE_ID']).replace('-', '')

        ws.cell(row=row_num, column=1, value=amount)
        cell_b = ws.cell(row=row_num, column=2, value=EXPRESSPAY_ACCOUNT)
        cell_b.number_format = '0'
        cell_b.alignment = center_align
        cell_c = ws.cell(row=row_num, column=3, value=EXPRESSPAY_NAME)
        cell_c.alignment = center_align
        cell_d = ws.cell(row=row_num, column=4, value=customer_account)
        cell_d.number_format = '0'
        cell_d.alignment = center_align
        cell_e = ws.cell(row=row_num, column=5, value=name)
        cell_e.alignment = center_align
        cell_f = ws.cell(row=row_num, column=6, value=narration)
        cell_f.alignment = center_align
        cell_g = ws.cell(row=row_num, column=7, value=session_id)
        cell_g.alignment = center_align
        row_num += 1

    extract_count = row_num - 2

    # Append BP Export failures (if any)
    bp_count = 0
    if bp_matched_df is not None and not bp_matched_df.empty:
        bp_df = bp_matched_df.copy()
        if bp_df['Amount'].dtype == 'object':
            bp_df['_amount'] = bp_df['Amount'].str.replace(',', '').astype(float)
        else:
            bp_df['_amount'] = bp_df['Amount'].astype(float)

        bp_df = bp_df.sort_values('_amount', ascending=True).reset_index(drop=True)

        for idx in bp_df.index:
            amount = bp_df.loc[idx, '_amount']
            if amount == int(amount):
                amount = int(amount)

            customer_account = bp_df.loc[idx, 'Acc Num1']
            raw_name = str(bp_df.loc[idx, 'Account1 Name']).strip()
            name = raw_name
            for title in ['MR ', 'MRS ', 'MS ', 'MISS ', 'DR. ', 'DR ']:
                if name.upper().startswith(title):
                    name = name[len(title):]
                    break

            # BP Export uses 'timestamp' for date and 'client_ref' for trace ID
            txn_date = str(bp_df.loc[idx, 'timestamp'])
            narration = build_narration(txn_date)
            session_id = str(bp_df.loc[idx, 'client_ref']).replace('-', '')

            ws.cell(row=row_num, column=1, value=amount)
            cell_b = ws.cell(row=row_num, column=2, value=EXPRESSPAY_ACCOUNT)
            cell_b.number_format = '0'
            cell_b.alignment = center_align
            cell_c = ws.cell(row=row_num, column=3, value=EXPRESSPAY_NAME)
            cell_c.alignment = center_align
            cell_d = ws.cell(row=row_num, column=4, value=customer_account)
            cell_d.number_format = '0'
            cell_d.alignment = center_align
            cell_e = ws.cell(row=row_num, column=5, value=name)
            cell_e.alignment = center_align
            cell_f = ws.cell(row=row_num, column=6, value=narration)
            cell_f.alignment = center_align
            cell_g = ws.cell(row=row_num, column=7, value=session_id)
            cell_g.alignment = center_align
            row_num += 1

        bp_count = len(bp_df)
        print(f"  ExpressPay failures appended: {bp_count}")

    # Adjust column widths
    col_widths = {'A': 10, 'B': 18, 'C': 28, 'D': 18, 'E': 35, 'F': 35, 'G': 38}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    wb.save(output_path)
    total_count = extract_count + bp_count
    print(f"  Reversal file saved with {total_count} transactions ({extract_count} extract + {bp_count} ExpressPay)")
    return total_count


def load_bp_export(file_path):
    """Load BP Export CSV file from ExpressPay."""
    print(f"Loading ExpressPay file: {file_path}")
    df = pd.read_csv(file_path)
    # Only keep DEBIT rows (actual transactions, not CREDIT commission rows)
    df = df[df['action'] == 'DEBIT'].copy()
    print(f"  Total transactions (DEBIT rows): {len(df)}")
    return df


def filter_bp_export_failures(bp_df, existing_trace_ids=None):
    """
    Filter BP Export for failed transactions (status != 'Successful').
    Excludes any trace IDs already flagged as failures from the FNB extract to avoid duplicates.
    """
    print("\nFiltering ExpressPay failures...")

    failed = bp_df[bp_df['status'] != 'Successful'].copy()
    print(f"  Failed transactions in ExpressPay: {len(failed)}")

    if len(failed) == 0:
        print("  No failures found in ExpressPay.")
        return pd.DataFrame()

    # Remove duplicates within bp-export itself
    before_dedup = len(failed)
    failed = failed.drop_duplicates(subset='client_ref', keep='first')
    dupes_removed = before_dedup - len(failed)
    if dupes_removed > 0:
        print(f"  ExpressPay internal duplicates removed: {dupes_removed}")

    # Remove any that overlap with FNB extract failures (already being reversed)
    if existing_trace_ids is not None and len(existing_trace_ids) > 0:
        before_overlap = len(failed)
        failed = failed[~failed['client_ref'].isin(existing_trace_ids)]
        overlap_removed = before_overlap - len(failed)
        if overlap_removed > 0:
            print(f"  Already in extract failures (duplicates removed): {overlap_removed}")

    print(f"  Unique ExpressPay failures to reverse: {len(failed)}")
    return failed


def match_bp_export_transactions(bp_failed_df, pay_pen_df):
    """
    Match BP Export failed transactions with PAY_PEN_PROCESS using client_ref = Narration.
    """
    print("\nMatching ExpressPay failures against PAY_PEN...")

    pay_pen_df = pay_pen_df.copy()
    pay_pen_df['Narration'] = pay_pen_df['Narration'].astype(str).str.strip()
    bp_failed_df = bp_failed_df.copy()
    bp_failed_df['client_ref'] = bp_failed_df['client_ref'].astype(str).str.strip()

    matched = pd.merge(
        bp_failed_df,
        pay_pen_df,
        left_on='client_ref',
        right_on='Narration',
        how='inner'
    )

    print(f"  Matched ExpressPay failures: {len(matched)}")

    unmatched_count = len(bp_failed_df) - len(matched)
    if unmatched_count > 0:
        unmatched_ids = set(bp_failed_df['client_ref']) - set(matched['client_ref'])
        print(f"  WARNING: {unmatched_count} ExpressPay failures not found in PAY_PEN_PROCESS")
        for uid in list(unmatched_ids)[:5]:
            print(f"    - {uid}")

    return matched


def extract_date_from_filename(filename):
    """Extract the date from an ExpressPay filename (YYYYMMDD at end)."""
    match = re.search(r'(\d{8})\.\w+$', filename)
    if match:
        return match.group(1)
    return None


def sort_expresspay_files_by_date(file_list):
    """Sort ExpressPay files by the date embedded in their filenames."""
    dated_files = []
    for f in file_list:
        date_str = extract_date_from_filename(f)
        if date_str:
            dated_files.append((date_str, f))
        else:
            dated_files.append(('00000000', f))
    dated_files.sort(key=lambda x: x[0])
    return [f for _, f in dated_files]


def process_single_extract(expresspay_file, pay_pen_df):
    """Process a single ExpressPay extract file and return matched transactions."""
    basename = os.path.basename(expresspay_file)
    print(f"\n{'─' * 60}")
    print(f"Processing: {basename}")
    print(f"{'─' * 60}")

    expresspay_df = load_expresspay_extract(expresspay_file)
    failed_df = filter_failed_transactions(expresspay_df)

    print(f"  Summary: {len(expresspay_df)} total | "
          f"{len(expresspay_df[expresspay_df['RESULT_MSG'] == 'Success'])} success | "
          f"{len(failed_df)} failed (timeout) | "
          f"{len(expresspay_df[expresspay_df['RESULT_MSG'] == 'RETRY_LIMIT_REACHED'])} retry (excluded)")

    if len(failed_df) == 0:
        print("  No failed transactions requiring reversal — skipping.")
        return pd.DataFrame()

    matched_df = match_transactions(failed_df, pay_pen_df)

    if len(matched_df) == 0:
        print("  No matches found — skipping.")
        return pd.DataFrame()

    return matched_df


def save_daily_stats(script_dir, expresspay_df, failed_df, total_reversals, total_amount, bp_export_failures=0):
    """Save today's processing stats to a persistent JSON log."""
    dashboard_dir = os.path.join(script_dir, 'dashboard')
    if not os.path.exists(dashboard_dir):
        os.makedirs(dashboard_dir)
    log_path = os.path.join(dashboard_dir, STATS_LOG_FILE)

    # Load existing log
    if os.path.exists(log_path):
        with open(log_path, 'r') as f:
            stats_log = json.load(f)
    else:
        stats_log = []

    today = datetime.now().strftime('%Y-%m-%d')

    # Count failure types
    failure_counts = {}
    for ftype in ['FULFILMENT_TIMEOUT', 'Third party Timeout', 'Invalid Request', 'Insufficient Funds']:
        failure_counts[ftype] = int(len(failed_df[failed_df['RESULT_MSG'] == ftype]))

    total_txns = int(len(expresspay_df))
    total_success = int(len(expresspay_df[expresspay_df['RESULT_MSG'] == 'Success']))
    total_failed = int(len(failed_df))
    retry_count = int(len(expresspay_df[expresspay_df['RESULT_MSG'] == 'RETRY_LIMIT_REACHED']))

    entry = {
        'date': today,
        'total_transactions': total_txns,
        'successful': total_success,
        'total_failed': total_failed,
        'failure_rate': round((total_failed / total_txns * 100) if total_txns > 0 else 0, 2),
        'failure_types': failure_counts,
        'retry_excluded': retry_count,
        'expresspay_failures': int(bp_export_failures),
        'reversals_generated': int(total_reversals),
        'reversal_amount': round(float(total_amount), 2)
    }

    # Replace existing entry for today or append
    stats_log = [s for s in stats_log if s['date'] != today]
    stats_log.append(entry)
    stats_log.sort(key=lambda x: x['date'])

    with open(log_path, 'w') as f:
        json.dump(stats_log, f, indent=2)

    return stats_log


def build_dashboard(script_dir, stats_log):
    """Build an Excel dashboard with charts from the accumulated stats log."""
    dashboard_dir = os.path.join(script_dir, 'dashboard')
    if not os.path.exists(dashboard_dir):
        os.makedirs(dashboard_dir)
    dashboard_path = os.path.join(dashboard_dir, 'Reversal_Dashboard.xlsx')
    print(f"\nBuilding dashboard: {dashboard_path}")

    wb = Workbook()

    # ── Colour palette ──
    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
    SUBTITLE_FONT = Font(bold=True, size=11, color="1F4E79")
    THIN_BORDER = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    center = Alignment(horizontal='center', vertical='center')
    ALT_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")

    # ══════════════════════════════════════════════════════════
    # SHEET 1: Daily Summary Table
    # ══════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Daily Summary"
    ws.sheet_properties.tabColor = "1F4E79"

    # Title
    ws.merge_cells('A1:K1')
    title_cell = ws.cell(row=1, column=1, value="Airtime & Bundles Reversal Dashboard")
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:K2')
    ws.cell(row=2, column=1, value=f"Last updated: {datetime.now().strftime('%d %b %Y %H:%M')}").font = Font(italic=True, color="808080")
    ws.cell(row=2, column=1).alignment = Alignment(horizontal='center')

    # Headers (row 4)
    headers = [
        "Date", "Total Txns", "Successful", "Failed",
        "Failure Rate %", "Fulfilment Timeout", "3rd Party Timeout",
        "Invalid Request", "Insufficient Funds", "Reversals", "Reversal Amount (GHS)"
    ]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = center
        cell.border = THIN_BORDER

    # Data rows
    for row_idx, entry in enumerate(stats_log, 5):
        ft = entry.get('failure_types', {})
        values = [
            entry['date'],
            entry['total_transactions'],
            entry['successful'],
            entry['total_failed'],
            entry['failure_rate'],
            ft.get('FULFILMENT_TIMEOUT', 0),
            ft.get('Third party Timeout', 0),
            ft.get('Invalid Request', 0),
            ft.get('Insufficient Funds', 0),
            entry['reversals_generated'],
            entry['reversal_amount']
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = center
            cell.border = THIN_BORDER
            if (row_idx - 5) % 2 == 1:
                cell.fill = ALT_FILL
            # Format percentage
            if col_idx == 5:
                cell.number_format = '0.00"%"'
            # Format currency
            if col_idx == 11:
                cell.number_format = '#,##0.00'

    col_widths = [14, 12, 12, 10, 14, 18, 18, 16, 18, 12, 22]
    for i, w in enumerate(col_widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    data_end_row = 4 + len(stats_log)

    # ══════════════════════════════════════════════════════════
    # SHEET 2: Charts (matplotlib images)
    # ══════════════════════════════════════════════════════════
    ws_charts = wb.create_sheet("Charts")
    ws_charts.sheet_properties.tabColor = "2E75B6"

    ws_charts.merge_cells('A1:R1')
    ws_charts.cell(row=1, column=1, value="Reversal Trends & Breakdown").font = TITLE_FONT
    ws_charts.cell(row=1, column=1).alignment = Alignment(horizontal='center')
    ws_charts.row_dimensions[1].height = 30

    if len(stats_log) >= 1:
        # Extract data for charts
        dates = [e['date'] for e in stats_log]
        short_dates = [d[5:] for d in dates]  # MM-DD for cleaner x-axis
        successful = [e['successful'] for e in stats_log]
        failed = [e['total_failed'] for e in stats_log]
        fail_rates = [e['failure_rate'] for e in stats_log]
        rev_amounts = [e['reversal_amount'] for e in stats_log]
        latest = stats_log[-1]
        ft = latest.get('failure_types', {})

        # Professional style settings
        plt.rcParams.update({
            'font.family': 'Segoe UI',
            'font.size': 10,
            'axes.spines.top': False,
            'axes.spines.right': False,
            'axes.grid': True,
            'grid.alpha': 0.3,
            'grid.linestyle': '--',
            'figure.facecolor': 'white',
            'axes.facecolor': '#FAFBFC',
        })

        BLUE = '#2E75B6'
        RED = '#E74C3C'
        ORANGE = '#E67E22'
        GREEN = '#27AE60'
        PURPLE = '#8E44AD'
        GOLD = '#F39C12'

        chart_paths = []

        # ── Chart 1: Daily Transaction Volume ──
        fig, ax = plt.subplots(figsize=(7, 4))
        x = range(len(dates))
        bar_w = 0.35
        bars1 = ax.bar([i - bar_w/2 for i in x], successful, bar_w, label='Successful', color=BLUE, edgecolor='white', linewidth=0.5)
        bars2 = ax.bar([i + bar_w/2 for i in x], failed, bar_w, label='Failed', color=RED, edgecolor='white', linewidth=0.5)
        for bar in bars1:
            ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.5, int(bar.get_height()), ha='center', va='bottom', fontsize=8, fontweight='bold', color=BLUE)
        for bar in bars2:
            ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.5, int(bar.get_height()), ha='center', va='bottom', fontsize=8, fontweight='bold', color=RED)
        ax.set_xticks(list(x))
        ax.set_xticklabels(short_dates, rotation=45, ha='right')
        ax.set_title('Daily Transaction Volume', fontsize=13, fontweight='bold', color='#1F4E79', pad=12)
        ax.set_ylabel('Count')
        ax.legend(frameon=True, fancybox=True, shadow=True, loc='upper right')
        ax.yaxis.set_major_locator(mticker.MaxNLocator(integer=True))
        fig.tight_layout()
        p1 = os.path.join(script_dir, '_chart_volume.png')
        fig.savefig(p1, dpi=150, bbox_inches='tight')
        plt.close(fig)
        chart_paths.append(p1)

        # ── Chart 2: Failure Rate Trend ──
        fig, ax = plt.subplots(figsize=(7, 4))
        ax.plot(short_dates, fail_rates, color=ORANGE, marker='o', linewidth=2.5, markersize=8, markerfacecolor='white', markeredgewidth=2, markeredgecolor=ORANGE)
        for i, rate in enumerate(fail_rates):
            ax.annotate(f'{rate:.1f}%', (short_dates[i], rate), textcoords="offset points", xytext=(0, 12), ha='center', fontsize=9, fontweight='bold', color=ORANGE)
        ax.set_title('Failure Rate Trend', fontsize=13, fontweight='bold', color='#1F4E79', pad=12)
        ax.set_ylabel('Failure Rate (%)')
        ax.set_ylim(0, max(fail_rates) * 1.3 if fail_rates else 100)
        ax.fill_between(short_dates, fail_rates, alpha=0.1, color=ORANGE)
        fig.tight_layout()
        p2 = os.path.join(script_dir, '_chart_failrate.png')
        fig.savefig(p2, dpi=150, bbox_inches='tight')
        plt.close(fig)
        chart_paths.append(p2)

        # ── Chart 3: Failure Type Breakdown (Donut) ──
        fig, ax = plt.subplots(figsize=(7, 4))
        pie_labels = list(ft.keys())
        pie_values = list(ft.values())
        colours = [BLUE, RED, GOLD, GREEN, PURPLE]
        # Filter out zero values for cleaner pie
        non_zero = [(l, v, c) for l, v, c in zip(pie_labels, pie_values, colours) if v > 0]
        if non_zero:
            nl, nv, nc = zip(*non_zero)
        else:
            nl, nv, nc = pie_labels, pie_values, colours[:len(pie_labels)]
        wedges, texts, autotexts = ax.pie(
            nv, labels=nl, colors=nc, autopct='%1.1f%%',
            startangle=90, pctdistance=0.75, wedgeprops=dict(width=0.45, edgecolor='white', linewidth=2),
            textprops={'fontsize': 9}
        )
        for at in autotexts:
            at.set_fontweight('bold')
            at.set_fontsize(9)
        ax.set_title(f'Failure Type Breakdown — {latest["date"]}', fontsize=13, fontweight='bold', color='#1F4E79', pad=12)
        # Add center text
        ax.text(0, 0, f'{sum(nv)}\nTotal', ha='center', va='center', fontsize=14, fontweight='bold', color='#1F4E79')
        fig.tight_layout()
        p3 = os.path.join(script_dir, '_chart_failtype.png')
        fig.savefig(p3, dpi=150, bbox_inches='tight')
        plt.close(fig)
        chart_paths.append(p3)

        # ── Chart 4: Reversal Amount ──
        fig, ax = plt.subplots(figsize=(7, 4))
        bars = ax.bar(short_dates, rev_amounts, color=GREEN, edgecolor='white', linewidth=0.5, width=0.5)
        for bar in bars:
            ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 10, f'GHS {bar.get_height():,.0f}', ha='center', va='bottom', fontsize=9, fontweight='bold', color=GREEN)
        ax.set_title('Daily Reversal Amount', fontsize=13, fontweight='bold', color='#1F4E79', pad=12)
        ax.set_ylabel('Amount (GHS)')
        ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f'{x:,.0f}'))
        fig.tight_layout()
        p4 = os.path.join(script_dir, '_chart_revamount.png')
        fig.savefig(p4, dpi=150, bbox_inches='tight')
        plt.close(fig)
        chart_paths.append(p4)

        # Embed images into Charts sheet in a 2x2 grid with proper spacing
        img1 = XlImage(chart_paths[0])
        img1.width = 620
        img1.height = 400
        img1.anchor = 'A3'
        ws_charts.add_image(img1)

        img2 = XlImage(chart_paths[1])
        img2.width = 620
        img2.height = 400
        img2.anchor = 'J3'
        ws_charts.add_image(img2)

        img3 = XlImage(chart_paths[2])
        img3.width = 620
        img3.height = 400
        img3.anchor = 'A28'
        ws_charts.add_image(img3)

        img4 = XlImage(chart_paths[3])
        img4.width = 620
        img4.height = 400
        img4.anchor = 'J28'
        ws_charts.add_image(img4)

    # ══════════════════════════════════════════════════════════
    # SHEET 3: Today's Snapshot (KPI cards as styled cells)
    # ══════════════════════════════════════════════════════════
    ws_snap = wb.create_sheet("Today's Snapshot")
    ws_snap.sheet_properties.tabColor = "27AE60"

    latest = stats_log[-1]
    ft = latest.get('failure_types', {})

    ws_snap.merge_cells('A1:H1')
    ws_snap.cell(row=1, column=1, value=f"Snapshot — {latest['date']}").font = TITLE_FONT
    ws_snap.cell(row=1, column=1).alignment = Alignment(horizontal='center')
    ws_snap.row_dimensions[1].height = 30

    kpi_cards = [
        ("Total Transactions", latest['total_transactions'], "1F4E79"),
        ("Successful", latest['successful'], "27AE60"),
        ("Failed", latest['total_failed'], "E74C3C"),
        ("Failure Rate", f"{latest['failure_rate']}%", "E67E22"),
        ("Reversals Generated", latest['reversals_generated'], "2E75B6"),
        ("Reversal Amount", f"GHS {latest['reversal_amount']:,.2f}", "8E44AD"),
        ("Fulfilment Timeout", ft.get('FULFILMENT_TIMEOUT', 0), "3498DB"),
        ("3rd Party Timeout", ft.get('Third party Timeout', 0), "E74C3C"),
        ("Invalid Request", ft.get('Invalid Request', 0), "F39C12"),
        ("Insufficient Funds", ft.get('Insufficient Funds', 0), "E74C3C"),
    ]

    for i, (label, value, colour) in enumerate(kpi_cards):
        col = (i % 3) * 3 + 1
        row = 3 + (i // 3) * 3

        # Label cell
        label_cell = ws_snap.cell(row=row, column=col, value=label)
        label_cell.font = Font(bold=True, size=10, color="FFFFFF")
        label_cell.fill = PatternFill(start_color=colour, end_color=colour, fill_type="solid")
        label_cell.alignment = center
        ws_snap.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)

        # Value cell
        value_cell = ws_snap.cell(row=row + 1, column=col, value=value)
        value_cell.font = Font(bold=True, size=16, color=colour)
        value_cell.alignment = center
        ws_snap.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 1)

    for i in range(1, 10):
        ws_snap.column_dimensions[get_column_letter(i)].width = 16

    wb.save(dashboard_path)
    print(f"  Dashboard saved with {len(stats_log)} day(s) of data")

    # Clean up temporary chart images
    for tmp in ['_chart_volume.png', '_chart_failrate.png', '_chart_failtype.png', '_chart_revamount.png']:
        tmp_path = os.path.join(script_dir, tmp)
        if os.path.exists(tmp_path):
            os.remove(tmp_path)


def main():
    """Main function to run the reversal processor."""
    print("=" * 60)
    print("AIRTIME & BUNDLES REVERSAL PROCESSOR")
    print("=" * 60)
    print(f"Run Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()

    # Get the script directory
    script_dir = os.path.dirname(os.path.abspath(__file__))

    # Find input files
    input_dir = os.path.join(script_dir, 'input')
    if not os.path.exists(input_dir):
        os.makedirs(input_dir)
    expresspay_files = [f for f in os.listdir(input_dir)
                        if 'EXPRESSPAY_EXTRACT' in f and f.endswith('.csv')]
    pay_pen_files = [f for f in os.listdir(input_dir)
                     if 'PAY_PEN_PROCESS' in f and f.endswith('.csv')]
    bp_export_files = [f for f in os.listdir(input_dir)
                       if f.startswith('bp-export') and f.endswith('.csv')]

    if not expresspay_files:
        print("ERROR: No ExpressPay Extract file found!")
        print("Place a CSV with 'EXPRESSPAY_EXTRACT' in the name in the input/ folder.")
        return

    if not pay_pen_files:
        print("ERROR: No PAY_PEN_PROCESS file found!")
        print("Place a CSV with 'PAY_PEN_PROCESS' in the name in the input/ folder.")
        return

    # Sort ExpressPay files by date (oldest first)
    expresspay_files = sort_expresspay_files_by_date(expresspay_files)
    pay_pen_file = os.path.join(input_dir, sorted(pay_pen_files)[-1])

    print(f"Found {len(expresspay_files)} ExpressPay extract file(s) — processing in date order:")
    for f in expresspay_files:
        print(f"  • {f}")
    if bp_export_files:
        print(f"\nExpressPay file(s): {len(bp_export_files)}")
        for f in bp_export_files:
            print(f"  • {f}")
    else:
        print("\nNo ExpressPay file found — skipping vendor-side failure check.")
    print(f"\nPAY_PEN file: {os.path.basename(pay_pen_file)}")

    # Load PAY_PEN once (shared across all extracts)
    pay_pen_df = load_pay_pen_process(pay_pen_file)

    # Process each extract file in date order and collect all matched transactions
    all_matched = []
    for filename in expresspay_files:
        filepath = os.path.join(input_dir, filename)
        matched_df = process_single_extract(filepath, pay_pen_df)
        if not matched_df.empty:
            all_matched.append(matched_df)

    if not all_matched:
        print("\nNo reversals to generate from extract files.")
        combined_df = pd.DataFrame()
    else:
        combined_df = pd.concat(all_matched, ignore_index=True)

    # ── Process BP Export files (vendor-side failures) ──
    bp_matched_df = pd.DataFrame()
    if bp_export_files:
        # Collect trace IDs already flagged from extract failures to avoid duplicates
        extract_trace_ids = set()
        if not combined_df.empty:
            extract_trace_ids = set(combined_df['TRACE_ID'].astype(str).str.strip())

        all_bp_matched = []
        for bp_file in bp_export_files:
            bp_path = os.path.join(input_dir, bp_file)
            print(f"\n{'─' * 60}")
            print(f"Processing ExpressPay: {bp_file}")
            print(f"{'─' * 60}")
            bp_df = load_bp_export(bp_path)
            bp_failed = filter_bp_export_failures(bp_df, existing_trace_ids=extract_trace_ids)

            if not bp_failed.empty:
                bp_matched = match_bp_export_transactions(bp_failed, pay_pen_df)
                if not bp_matched.empty:
                    all_bp_matched.append(bp_matched)
                    # Add these trace IDs to prevent duplicates across multiple bp-export files
                    extract_trace_ids.update(bp_matched['client_ref'].astype(str).str.strip())

        if all_bp_matched:
            bp_matched_df = pd.concat(all_bp_matched, ignore_index=True)

    # Check if we have anything to reverse at all
    if combined_df.empty and bp_matched_df.empty:
        print("\nNo reversals to generate from any source!")
        return

    # Generate single output file with today's date
    output_dir = os.path.join(script_dir, 'output')
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    today = datetime.now().strftime('%d%m%Y')
    output_file = os.path.join(output_dir, f'AirtimeBundles_reversal_{today}.xlsx')

    # If no extract failures but bp-export failures exist, create with empty extract df
    if combined_df.empty:
        combined_df = pd.DataFrame(columns=['Amount', 'Acc Num1', 'Account1 Name', 'TXN_DATE', 'TRACE_ID'])

    total_reversals = create_reversal_file(combined_df, output_file, bp_matched_df=bp_matched_df)

    # Calculate total amount
    total_amount = 0
    if not combined_df.empty and 'Amount' in combined_df.columns and len(combined_df) > 0:
        if combined_df['Amount'].dtype == 'object':
            total_amount += combined_df['Amount'].str.replace(',', '').astype(float).sum()
        else:
            total_amount += combined_df['Amount'].astype(float).sum()
    if not bp_matched_df.empty and 'Amount' in bp_matched_df.columns:
        if bp_matched_df['Amount'].dtype == 'object':
            total_amount += bp_matched_df['Amount'].str.replace(',', '').astype(float).sum()
        else:
            total_amount += bp_matched_df['Amount'].astype(float).sum()

    # Count ExpressPay failures for stats
    bp_export_failures = len(bp_matched_df) if not bp_matched_df.empty else 0

    # Combine all ExpressPay data for stats
    all_expresspay = []
    for filename in expresspay_files:
        filepath = os.path.join(input_dir, filename)
        all_expresspay.append(pd.read_csv(filepath))
    combined_expresspay = pd.concat(all_expresspay, ignore_index=True)

    # Rebuild combined failed for stats (unique truly-failed across all files)
    combined_failed = filter_failed_transactions(combined_expresspay)

    # Save stats and build dashboard
    stats_log = save_daily_stats(script_dir, combined_expresspay, combined_failed, total_reversals, total_amount, bp_export_failures)
    build_dashboard(script_dir, stats_log)

    # Print summary
    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)
    print(f"Files processed:                  {len(expresspay_files)} extract + {len(bp_export_files)} ExpressPay")
    print(f"Extract reversals:                {total_reversals - bp_export_failures}")
    print(f"ExpressPay reversals:             {bp_export_failures}")
    print(f"Total reversals generated:        {total_reversals}")
    print(f"Total reversal amount:            GHS {total_amount:,.2f}")
    print(f"Output file: {output_file}")
    print("=" * 60)


if __name__ == "__main__":
    main()
